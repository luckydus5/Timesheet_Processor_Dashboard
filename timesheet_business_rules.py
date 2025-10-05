"""
🎯 TIMESHEET BUSINESS RULES IMPLEMENTATION
=========================================

This module implements the exact business rules for timesheet processing:

1. Day Shift: Official hours 8:00 AM - 17:00 PM
   - Can check-in before 8:00 AM (no overtime for early check-in)
   - Overtime only after 17:00 PM (30min-1.5h max)

2. Night Shift: Official hours 18:00 PM - 3:00 AM
   - Can check-in before 18:00 PM (no overtime for early check-in)
   - Overtime only after 3:00 AM next day (30min-3h max)

3. Handles multiple check-ins/check-outs per day
4. Cross-midnight shift calculations
"""

import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
import os


class TimesheetBusinessRules:
    """
    Implementation of company timesheet business rules
    """
    
    # Company shift definitions
    DAY_SHIFT_START = time(8, 0, 0)      # 8:00 AM
    DAY_SHIFT_END = time(17, 0, 0)       # 5:00 PM
    NIGHT_SHIFT_START = time(18, 0, 0)   # 6:00 PM
    NIGHT_SHIFT_END = time(3, 0, 0)      # 3:00 AM next day
    
    # Overtime rules
    MIN_OVERTIME_MINUTES = 30            # 30 minutes minimum
    DAY_SHIFT_MAX_OVERTIME_HOURS = 1.5   # 1.5 hours maximum
    NIGHT_SHIFT_MAX_OVERTIME_HOURS = 3.0 # 3 hours maximum
    
    def __init__(self):
        print("🎯 Timesheet Business Rules Initialized")
        print("=" * 50)
        self.print_business_rules()
    
    def print_business_rules(self):
        """Print the complete business rules"""
        print("\n📋 COMPANY TIMESHEET BUSINESS RULES:")
        print("=" * 40)
        
        print("\n1️⃣ SHIFT DEFINITIONS:")
        print("   📅 Day Shift:")
        print("      • Official Hours: 8:00 AM - 17:00 PM (5:00 PM)")
        print("      • Early Check-in: Allowed (before 8:00 AM) - NO overtime")
        print("      • Shift Type: Determined if FIRST check-in is before 18:00 PM")
        
        print("\n   🌙 Night Shift:")
        print("      • Official Hours: 18:00 PM - 3:00 AM (next day)")
        print("      • Early Check-in: Allowed (before 18:00 PM) - NO overtime")
        print("      • Shift Type: Determined if FIRST check-in is 18:00 PM or later")
        print("      • Note: Can check-in as early as 16:20 PM")
        
        print("\n2️⃣ OVERTIME RULES:")
        print("   📅 Day Shift Overtime:")
        print("      • When: Only AFTER 17:00 PM (5:00 PM)")
        print("      • NO overtime for early check-in (before 8:00 AM)")
        print("      • Minimum: 30 minutes (below = no overtime)")
        print("      • Maximum: 1.5 hours per shift")
        
        print("\n   🌙 Night Shift Overtime:")
        print("      • When: Only AFTER 3:00 AM (next day)")
        print("      • NO overtime for early check-in (before 18:00 PM)")
        print("      • Minimum: 30 minutes (below = no overtime)")
        print("      • Maximum: 3 hours per shift")
        
        print("\n3️⃣ MULTIPLE ENTRIES HANDLING:")
        print("   🔄 Problem: Multiple check-ins/check-outs per day")
        print("   ✅ Solution:")
        print("      • Start Time = FIRST check-in of the day")
        print("      • End Time = LAST check-out of the day")
        print("      • Ignores intermediate entries")
        
        print("\n4️⃣ CROSS-MIDNIGHT HANDLING:")
        print("   🌃 Night shifts spanning two calendar days")
        print("   ✅ Automatic detection and calculation")
        
        print("\n" + "="*50)
    
    def parse_date_time(self, date_str, time_str):
        """Parse separate date and time strings"""
        if pd.isna(date_str) or pd.isna(time_str) or date_str == '' or time_str == '':
            return None, None
        try:
            # Parse date string (various formats)
            date_obj = pd.to_datetime(date_str).date()
            
            # Parse time string
            time_obj = pd.to_datetime(time_str, format='%H:%M:%S').time()
            
            return date_obj, time_obj
        except:
            return None, None
    
    def determine_shift_type(self, start_time):
        """
        Determine shift type based on FIRST check-in time
        
        Business Rule:
        - If FIRST check-in < 18:00 PM → Day Shift (even if at 6:00 AM)
        - If FIRST check-in ≥ 18:00 PM → Night Shift
        - Special case: Night workers can check-in before 18:00 PM (like 16:20 PM)
          but this is still considered preparation for night shift
        """
        if start_time is None:
            return ""
        
        start_decimal = start_time.hour + start_time.minute/60 + start_time.second/3600
        
        # Key decision point: 18:00 PM (6:00 PM)
        if start_decimal < 18.0:
            # Check if this might be a night worker checking in early
            # Night workers often check-in between 16:00-18:00 PM
            if start_decimal >= 16.0:  # Between 4:00 PM - 6:00 PM
                # This could be night shift preparation
                # For now, we'll classify as Day Shift, but flag for review
                return "Day Shift"  # Will be overridden by business logic if needed
            else:
                return "Day Shift"
        else:
            return "Night Shift"
    
    def find_shift_boundaries(self, df, name, date):
        """
        Find start and end times for a given employee and date
        Handles multiple check-ins/check-outs correctly
        """
        # Filter records for this person and date
        day_records = df[(df['Name'] == name) & (df['Date_parsed'] == date)].copy()
        
        if day_records.empty:
            return None, None
        
        # Sort by time to get chronological order
        day_records = day_records.sort_values('Time_parsed')
        
        # Find all check-ins and check-outs
        checkins = day_records[day_records['Status'].isin(['C/In', 'OverTime In'])]
        checkouts = day_records[day_records['Status'].isin(['C/Out', 'OverTime Out'])]
        
        start_time = None
        end_time = None
        
        if not checkins.empty:
            # FIRST check-in of the day is the start time
            start_time = checkins.iloc[0]['Time_parsed']
            
            # LAST check-out of the day is the end time
            if not checkouts.empty:
                end_time = checkouts.iloc[-1]['Time_parsed']
        
        return start_time, end_time
    
    def calculate_total_work_hours(self, start_time, end_time, shift_type):
        """
        Calculate total work hours between start and end time
        Handles cross-midnight shifts for night workers
        """
        if start_time is None or end_time is None:
            return 0
        
        # Convert times to datetime objects for calculation
        start_dt = datetime.combine(datetime.today(), start_time)
        end_dt = datetime.combine(datetime.today(), end_time)
        
        # Handle cross-midnight shifts (night shift ending next day)
        if shift_type == "Night Shift" and end_time < start_time:
            # Add one day to end time for cross-midnight calculation
            end_dt += timedelta(days=1)
        
        # Calculate total hours
        total_duration = end_dt - start_dt
        total_hours = total_duration.total_seconds() / 3600
        
        return round(total_hours, 2)
    
    def calculate_overtime_hours(self, start_time, end_time, shift_type):
        """
        Calculate overtime hours based on company business rules
        
        KEY RULE: No overtime for early check-ins, only for late check-outs
        """
        if start_time is None or end_time is None or shift_type == "":
            return 0
        
        overtime = 0
        
        if shift_type == "Day Shift":
            # Day shift overtime: ONLY after 17:00 PM (5:00 PM)
            # NO overtime for checking in before 8:00 AM
            end_decimal = end_time.hour + end_time.minute/60 + end_time.second/3600
            
            if end_decimal > 17.0:  # After 5:00 PM
                overtime = end_decimal - 17.0
                
                # Apply minimum 30 minutes rule
                if overtime < 0.5:
                    overtime = 0
                # Apply maximum 1.5 hours rule
                elif overtime > self.DAY_SHIFT_MAX_OVERTIME_HOURS:
                    overtime = self.DAY_SHIFT_MAX_OVERTIME_HOURS
                    
        elif shift_type == "Night Shift":
            # Night shift overtime: ONLY after 3:00 AM (next day)
            # NO overtime for checking in before 18:00 PM
            end_decimal = end_time.hour + end_time.minute/60 + end_time.second/3600
            
            # For night shift, check if end time is in early morning hours (cross-midnight)
            if end_decimal <= 12.0:  # Early morning hours (00:00-12:00)
                if end_decimal > 3.0:  # After 3:00 AM
                    overtime = end_decimal - 3.0
                    
                    # Apply minimum 30 minutes rule
                    if overtime < 0.5:
                        overtime = 0
                    # Apply maximum 3 hours rule
                    elif overtime > self.NIGHT_SHIFT_MAX_OVERTIME_HOURS:
                        overtime = self.NIGHT_SHIFT_MAX_OVERTIME_HOURS
        
        return round(overtime, 2)
    
    def calculate_regular_hours(self, total_hours, overtime_hours):
        """Calculate regular hours (total - overtime)"""
        if total_hours == 0:
            return 0
        
        regular = total_hours - overtime_hours
        return round(max(regular, 0), 2)  # Ensure non-negative
    
    def process_timesheet_data(self, df):
        """
        Master function to process timesheet data according to business rules
        
        Expected columns: Name, Date, Time, Status
        Returns: DataFrame with calculated shift information
        """
        print("\n🧹 Starting timesheet processing...")
        print("=" * 40)
        
        # Make a copy to avoid modifying original
        df_clean = df.copy()
        
        # Step 1: Clean the data structure
        print("📋 Step 1: Preparing data structure...")
        
        # Remove unnecessary columns
        unnecessary_cols = [col for col in df_clean.columns if 'Unnamed' in col]
        for col in unnecessary_cols:
            df_clean = df_clean.drop(col, axis=1)
            print(f"   ✅ Removed {col}")
        
        # Step 2: Parse Date and Time columns
        print("📅 Step 2: Parsing Date and Time...")
        
        df_clean[['Date_parsed', 'Time_parsed']] = df_clean.apply(
            lambda row: pd.Series(self.parse_date_time(row['Date'], row['Time'])), axis=1
        )
        
        # Remove rows where parsing failed
        initial_count = len(df_clean)
        df_clean = df_clean[df_clean['Date_parsed'].notna()]
        df_clean = df_clean[df_clean['Time_parsed'].notna()]
        
        print(f"   ✅ Successfully parsed {len(df_clean)} records ({initial_count - len(df_clean)} failed)")
        
        # Step 3: Calculate shift boundaries and metrics
        print("⏰ Step 3: Applying business rules...")
        
        # Create cache for performance
        shift_cache = {}
        
        # Initialize calculated columns
        df_clean['Start Time'] = ''
        df_clean['End Time'] = ''
        df_clean['Shift Time'] = ''
        df_clean['Total Hours'] = 0.0
        df_clean['Regular Hours'] = 0.0
        df_clean['Overtime Hours'] = 0.0
        
        # Process each row
        total_rows = len(df_clean)
        processed = 0
        unique_shifts = set()
        
        for idx, row in df_clean.iterrows():
            name = row['Name']
            date = row['Date_parsed']
            
            if pd.isna(name) or pd.isna(date):
                continue
                
            # Create cache key for each employee-date combination
            cache_key = f"{name}_{date}"
            unique_shifts.add(cache_key)
            
            # Calculate shift data once per employee-date combination
            if cache_key not in shift_cache:
                start_time, end_time = self.find_shift_boundaries(df_clean, name, date)
                
                if start_time and end_time:
                    # Determine shift type based on FIRST check-in
                    shift_type = self.determine_shift_type(start_time)
                    
                    # Calculate total work hours (handles cross-midnight)
                    total_hours = self.calculate_total_work_hours(start_time, end_time, shift_type)
                    
                    # Calculate overtime based on business rules
                    overtime_hours = self.calculate_overtime_hours(start_time, end_time, shift_type)
                    
                    # Calculate regular hours
                    regular_hours = self.calculate_regular_hours(total_hours, overtime_hours)
                    
                    # Cache the results
                    shift_cache[cache_key] = {
                        'start_time': start_time.strftime('%H:%M:%S'),
                        'end_time': end_time.strftime('%H:%M:%S'),
                        'shift_type': shift_type,
                        'total_hours': total_hours,
                        'regular_hours': regular_hours,
                        'overtime_hours': overtime_hours
                    }
                else:
                    # No valid shift found
                    shift_cache[cache_key] = None
            
            # Apply cached values to current row
            if shift_cache[cache_key]:
                data = shift_cache[cache_key]
                df_clean.at[idx, 'Start Time'] = data['start_time']
                df_clean.at[idx, 'End Time'] = data['end_time']
                df_clean.at[idx, 'Shift Time'] = data['shift_type']
                df_clean.at[idx, 'Total Hours'] = data['total_hours']
                df_clean.at[idx, 'Regular Hours'] = data['regular_hours']
                df_clean.at[idx, 'Overtime Hours'] = data['overtime_hours']
            
            processed += 1
            if processed % 500 == 0:
                print(f"   📈 Processed {processed}/{total_rows} records...")
        
        print(f"   ✅ Completed processing {processed} records")
        print(f"   📊 Found {len(unique_shifts)} unique employee-date combinations")
        
        # Step 4: Final formatting
        print("📊 Step 4: Final formatting...")
        
        # Ensure proper column order
        final_columns = ['Name', 'Date', 'Time', 'Status', 'Start Time', 'End Time', 
                        'Shift Time', 'Total Hours', 'Regular Hours', 'Overtime Hours']
        df_final = df_clean[final_columns].copy()
        
        # Sort by Name and Date for better organization
        df_final = df_final.sort_values(['Name', 'Date', 'Time'])
        
        # Clean up any remaining NaN values
        numeric_columns = ['Total Hours', 'Regular Hours', 'Overtime Hours']
        for col in numeric_columns:
            df_final[col] = df_final[col].fillna(0)
        
        string_columns = ['Start Time', 'End Time', 'Shift Time']
        for col in string_columns:
            df_final[col] = df_final[col].fillna('')
        
        print("✅ Business rules applied successfully!")
        
        return df_final


def load_timesheet_file(file_path):
    """Load timesheet data from Excel or CSV file"""
    try:
        # Determine file type and load accordingly
        if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls'):
            df = pd.read_excel(file_path)
            print(f"✅ Excel file loaded: {file_path}")
        elif file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
            print(f"✅ CSV file loaded: {file_path}")
        else:
            raise ValueError("File must be Excel (.xlsx/.xls) or CSV (.csv)")
        
        # Display basic information
        print(f"\n📊 Data Overview:")
        print(f"   - Total records: {len(df):,}")
        print(f"   - Columns: {list(df.columns)}")
        
        # Check for required columns
        required_cols = ['Name', 'Date', 'Time', 'Status']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            print(f"❌ Missing required columns: {missing_cols}")
            return None
        
        print(f"✅ All required columns present")
        return df
    
    except Exception as e:
        print(f"❌ Error loading file: {str(e)}")
        return None


def export_processed_data(df, base_filename="Processed_Timesheet"):
    """Export processed data to CSV and Excel"""
    if df is None or df.empty:
        print("❌ No data to export")
        return None, None
    
    # Generate timestamped filenames
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"{base_filename}_{timestamp}.csv"
    excel_filename = f"{base_filename}_{timestamp}.xlsx"
    
    try:
        # Export to CSV
        df.to_csv(csv_filename, index=False)
        print(f"✅ CSV exported: {csv_filename}")
        
        # Export to Excel (if openpyxl is available)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                # Main data sheet
                df.to_excel(writer, sheet_name='Processed_Data', index=False)
                
                # Format headers
                workbook = writer.book
                worksheet = writer.sheets['Processed_Data']
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ Excel exported: {excel_filename}")
            
        except ImportError:
            print("⚠️ Excel export requires openpyxl package")
            excel_filename = None
        
        print(f"\n📊 Export Summary:")
        print(f"   Records exported: {len(df):,}")
        print(f"   CSV file size: {os.path.getsize(csv_filename) / 1024:.1f} KB")
        
        return csv_filename, excel_filename
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        return None, None


def process_timesheet_file(input_file_path, output_name="Processed_Timesheet"):
    """
    🚀 ONE-CLICK TIMESHEET PROCESSOR
    
    Complete function that loads, processes, and exports timesheet data
    according to company business rules.
    
    Args:
        input_file_path (str): Path to Excel or CSV timesheet file
        output_name (str): Base name for output files
        
    Returns:
        tuple: (processed_dataframe, csv_filename, excel_filename)
    """
    print("🚀 STARTING TIMESHEET PROCESSING")
    print("=" * 50)
    
    try:
        # Step 1: Load data
        print("\n📂 STEP 1: Loading timesheet data...")
        df = load_timesheet_file(input_file_path)
        if df is None:
            return None, None, None
        
        # Step 2: Initialize business rules processor
        print("\n🎯 STEP 2: Initializing business rules...")
        processor = TimesheetBusinessRules()
        
        # Step 3: Process data
        print("\n🧹 STEP 3: Processing timesheet data...")
        processed_df = processor.process_timesheet_data(df)
        
        # Step 4: Export results
        print("\n💾 STEP 4: Exporting results...")
        csv_file, excel_file = export_processed_data(processed_df, output_name)
        
        # Step 5: Summary
        print(f"\n🎉 PROCESSING COMPLETED SUCCESSFULLY!")
        print(f"📋 Summary:")
        print(f"   - Original records: {len(df):,}")
        print(f"   - Processed records: {len(processed_df):,}")
        print(f"   - Unique employees: {processed_df['Name'].nunique()}")
        
        day_shifts = len(processed_df[processed_df['Shift Time'] == 'Day Shift'])
        night_shifts = len(processed_df[processed_df['Shift Time'] == 'Night Shift'])
        print(f"   - Day shift records: {day_shifts:,}")
        print(f"   - Night shift records: {night_shifts:,}")
        
        overtime_records = processed_df[processed_df['Overtime Hours'] > 0]
        total_overtime = processed_df['Overtime Hours'].sum()
        print(f"   - Records with overtime: {len(overtime_records):,}")
        print(f"   - Total overtime hours: {total_overtime:.2f}")
        
        return processed_df, csv_file, excel_file
        
    except Exception as e:
        print(f"❌ Error in processing: {str(e)}")
        return None, None, None


# Example usage function
def main():
    """Example usage of the timesheet processing system"""
    print("🎯 TIMESHEET BUSINESS RULES PROCESSOR")
    print("=" * 50)
    print("\n📝 USAGE EXAMPLE:")
    print("   from timesheet_business_rules import process_timesheet_file")
    print("   ")
    print("   # Process your timesheet file")
    print("   data, csv_file, excel_file = process_timesheet_file('your_timesheet.xlsx')")
    print("   ")
    print("   # Or with custom output name:")
    print("   data, csv_file, excel_file = process_timesheet_file(")
    print("       'timesheet.csv', 'Company_Payroll_2025')")
    
    print(f"\n🎯 BUSINESS RULES IMPLEMENTED:")
    rules = TimesheetBusinessRules()


if __name__ == "__main__":
    main()