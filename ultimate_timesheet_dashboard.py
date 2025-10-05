#!/usr/bin/env python3
"""
🏆 ULTIMATE TIMESHEET PROCESSOR - Premium Edition
Professional Excel Enhancement System

Features:
- Preserves original Excel structure (NO SORTING!)
- Adds calculated columns without changing order
- Maintains all original names and data
- Premium visual design
- Professional Excel formatting
- Real-time processing feedback

Author: AI Assistant - Premium Edition
Created: October 2025
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
import os
from typing import Tuple, Optional, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Premium Page Configuration
st.set_page_config(
    page_title="🏆 Ultimate Timesheet Processor",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Premium CSS Styling
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    .main {
        font-family: 'Inter', sans-serif;
    }
    
    .premium-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    }
    
    .premium-title {
        font-size: 3.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .premium-subtitle {
        font-size: 1.3rem;
        font-weight: 300;
        opacity: 0.9;
    }
    
    .feature-card {
        background: linear-gradient(145deg, #f8f9ff 0%, #e8f2ff 100%);
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid #e1e8ff;
        margin-bottom: 1rem;
        box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        transition: transform 0.3s ease;
    }
    
    .feature-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    }
    
    .metric-premium {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        text-align: center;
        box-shadow: 0 5px 15px rgba(79, 172, 254, 0.3);
    }
    
    .metric-premium h1 {
        font-size: 2.5rem;
        margin: 0;
        font-weight: 700;
    }
    
    .metric-premium p {
        margin: 0;
        font-size: 1.1rem;
        opacity: 0.9;
    }
    
    .success-premium {
        background: linear-gradient(135deg, #4CAF50 0%, #81C784 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(76, 175, 80, 0.3);
    }
    
    .warning-premium {
        background: linear-gradient(135deg, #FF9800 0%, #FFB74D 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(255, 152, 0, 0.3);
    }
    
    .info-premium {
        background: linear-gradient(135deg, #2196F3 0%, #64B5F6 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(33, 150, 243, 0.3);
    }
    
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #f8f9ff 0%, #e8f2ff 100%);
    }
    
    .premium-button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.75rem 1.5rem;
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
    
    .premium-button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }
    
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 5px 15px rgba(0,0,0,0.1);
    }
    
    .progress-container {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        height: 6px;
        border-radius: 3px;
        overflow: hidden;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

class PremiumTimesheetProcessor:
    """Premium timesheet processor that preserves original structure"""
    
    def __init__(self):
        self.BASE_FOLDER = "/home/luckdus/Desktop/Data Cleaner"
    
    def parse_date_time(self, date_str, time_str):
        """Parse separate date and time strings"""
        if pd.isna(date_str) or pd.isna(time_str) or date_str == '' or time_str == '':
            return None, None
        try:
            date_obj = pd.to_datetime(date_str, dayfirst=True).date()
            time_obj = pd.to_datetime(time_str, format='%H:%M:%S').time()
            return date_obj, time_obj
        except:
            return None, None

    def load_timesheet_file_premium(self, uploaded_file) -> Optional[pd.DataFrame]:
        """Load timesheet data preserving original structure"""
        try:
            if uploaded_file.name.lower().endswith('.xlsx') or uploaded_file.name.lower().endswith('.xls'):
                df = pd.read_excel(uploaded_file)
                st.success("✅ Excel file loaded successfully!")
            elif uploaded_file.name.lower().endswith('.csv'):
                df = pd.read_csv(uploaded_file)
                st.success("✅ CSV file loaded successfully!")
            else:
                st.error("❌ File must be Excel (.xlsx/.xls) or CSV (.csv)")
                return None
            
            # Handle different file formats WITHOUT changing original structure
            original_df = df.copy()  # Keep original intact
            
            if 'Date/Time' in df.columns:
                st.info("🔄 Detected combined Date/Time column - processing...")
                df['DateTime_parsed'] = pd.to_datetime(df['Date/Time'], errors='coerce')
                df['Date'] = df['DateTime_parsed'].dt.strftime('%d/%m/%Y')
                df['Time'] = df['DateTime_parsed'].dt.strftime('%H:%M:%S')
            
            # Check for required columns
            required_cols = ['Name', 'Date', 'Time', 'Status']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                st.error(f"❌ Missing required columns: {missing_cols}")
                st.info(f"💡 Available columns: {list(df.columns)}")
                return None
            
            # Store original structure for preservation
            df._original_df = original_df
            df._original_index = original_df.index.copy()
            
            return df
            
        except Exception as e:
            st.error(f"❌ Error loading file: {str(e)}")
            return None

    def calculate_enhanced_data(self, df) -> pd.DataFrame:
        """Calculate enhanced data while preserving original structure"""
        
        # Create progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Make a working copy
        enhanced_df = df.copy()
        
        # Add processing columns while preserving original order
        enhanced_df['Date_parsed'] = None
        enhanced_df['Time_parsed'] = None
        enhanced_df['Employee_Date_Key'] = None
        enhanced_df['Consolidated_Start_Time'] = None
        enhanced_df['Consolidated_End_Time'] = None
        enhanced_df['Shift_Type'] = None
        enhanced_df['Total_Work_Hours'] = None
        enhanced_df['Regular_Hours'] = None
        enhanced_df['Overtime_Hours'] = None
        enhanced_df['Entry_Count_For_Day'] = None
        enhanced_df['Is_First_Entry'] = False
        enhanced_df['Is_Last_Entry'] = False
        
        status_text.text("🔄 Step 1: Parsing dates and times...")
        progress_bar.progress(0.2)
        
        # Parse dates and times
        for idx, row in enhanced_df.iterrows():
            date_parsed, time_parsed = self.parse_date_time(row['Date'], row['Time'])
            enhanced_df.at[idx, 'Date_parsed'] = date_parsed
            enhanced_df.at[idx, 'Time_parsed'] = time_parsed
            if date_parsed and time_parsed:
                enhanced_df.at[idx, 'Employee_Date_Key'] = f"{row['Name']}_{date_parsed}"
        
        status_text.text("🔄 Step 2: Analyzing employee patterns...")
        progress_bar.progress(0.4)
        
        # Group by employee and date to calculate consolidated data
        employee_date_groups = {}
        
        for idx, row in enhanced_df.iterrows():
            if pd.notna(row['Employee_Date_Key']):
                key = row['Employee_Date_Key']
                if key not in employee_date_groups:
                    employee_date_groups[key] = []
                employee_date_groups[key].append({
                    'index': idx,
                    'time': row['Time_parsed'],
                    'status': row['Status'],
                    'name': row['Name'],
                    'date': row['Date_parsed']
                })
        
        status_text.text("🔄 Step 3: Calculating consolidated times...")
        progress_bar.progress(0.6)
        
        # Calculate consolidated data for each employee-date combination
        for key, entries in employee_date_groups.items():
            # Sort entries by time
            entries.sort(key=lambda x: x['time'] if x['time'] else time.min)
            
            # Find first check-in and last check-out
            checkins = [e for e in entries if e['status'] in ['C/In', 'OverTime In']]
            checkouts = [e for e in entries if e['status'] in ['C/Out', 'OverTime Out']]
            
            start_time = checkins[0]['time'] if checkins else None
            end_time = checkouts[-1]['time'] if checkouts else None
            
            # Determine shift type
            shift_type = ""
            if start_time:
                start_decimal = start_time.hour + start_time.minute/60 + start_time.second/3600
                shift_type = "Day Shift" if start_decimal < 18.0 else "Night Shift"
            
            # Calculate hours
            total_hours = self.calculate_total_work_hours(start_time, end_time, shift_type)
            overtime_hours = self.calculate_overtime_hours(start_time, end_time, shift_type)
            regular_hours = total_hours - overtime_hours if total_hours > 0 else 0
            
            # Update all entries for this employee-date
            for i, entry in enumerate(entries):
                idx = entry['index']
                enhanced_df.at[idx, 'Consolidated_Start_Time'] = start_time.strftime('%H:%M:%S') if start_time else ""
                enhanced_df.at[idx, 'Consolidated_End_Time'] = end_time.strftime('%H:%M:%S') if end_time else ""
                enhanced_df.at[idx, 'Shift_Type'] = shift_type
                enhanced_df.at[idx, 'Total_Work_Hours'] = total_hours
                enhanced_df.at[idx, 'Regular_Hours'] = regular_hours
                enhanced_df.at[idx, 'Overtime_Hours'] = overtime_hours
                enhanced_df.at[idx, 'Entry_Count_For_Day'] = len(entries)
                enhanced_df.at[idx, 'Is_First_Entry'] = (i == 0)
                enhanced_df.at[idx, 'Is_Last_Entry'] = (i == len(entries) - 1)
        
        status_text.text("🔄 Step 4: Finalizing calculations...")
        progress_bar.progress(0.8)
        
        # Clean up helper columns
        enhanced_df = enhanced_df.drop(['Date_parsed', 'Time_parsed', 'Employee_Date_Key'], axis=1)
        
        progress_bar.progress(1.0)
        status_text.text("✅ Processing complete!")
        
        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
        
        return enhanced_df

    def calculate_total_work_hours(self, start_time, end_time, shift_type):
        """Calculate total work hours"""
        if start_time is None or end_time is None:
            return 0
        
        start_dt = datetime.combine(datetime.today(), start_time)
        end_dt = datetime.combine(datetime.today(), end_time)
        
        if shift_type == "Night Shift" and end_time < start_time:
            end_dt += timedelta(days=1)
        
        total_duration = end_dt - start_dt
        total_hours = total_duration.total_seconds() / 3600
        return round(total_hours, 2)

    def calculate_overtime_hours(self, start_time, end_time, shift_type):
        """Calculate overtime hours"""
        if start_time is None or end_time is None or shift_type == "":
            return 0
        
        overtime = 0
        
        if shift_type == "Day Shift":
            end_decimal = end_time.hour + end_time.minute/60 + end_time.second/3600
            if end_decimal > 17.0:
                overtime = end_decimal - 17.0
                if overtime < 0.5:
                    overtime = 0
                elif overtime > 1.5:
                    overtime = 1.5
                    
        elif shift_type == "Night Shift":
            end_decimal = end_time.hour + end_time.minute/60 + end_time.second/3600
            if end_decimal <= 12.0:
                if end_decimal > 3.0:
                    overtime = end_decimal - 3.0
                    if overtime < 0.5:
                        overtime = 0
                    elif overtime > 3.0:
                        overtime = 3.0
        
        return round(overtime, 2)

    def export_premium_excel(self, df, filename="Enhanced_Timesheet"):
        """Export with premium Excel formatting"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{filename}_{timestamp}.xlsx"
        
        # Create Excel file with premium formatting
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Enhanced_Timesheet', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Enhanced_Timesheet']
            
            # Premium styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            
            # Enhanced columns styling
            enhanced_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
            enhanced_font = Font(color="1B4F72", bold=True, size=10)
            
            # Apply header formatting
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Highlight enhanced columns
                if column_title in ['Consolidated_Start_Time', 'Consolidated_End_Time', 'Shift_Type', 
                                   'Total_Work_Hours', 'Regular_Hours', 'Overtime_Hours']:
                    for row_num in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = enhanced_fill
                        cell.font = enhanced_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create summary sheet
            summary_data = self.create_summary_data(df)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Processing_Summary', index=False)
            
            # Format summary sheet
            summary_sheet = writer.sheets['Processing_Summary']
            for col_num, column_title in enumerate(summary_df.columns, 1):
                cell = summary_sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        excel_data = output.getvalue()
        return excel_data, excel_filename

    def create_summary_data(self, df):
        """Create summary statistics"""
        unique_employees = df['Name'].nunique()
        total_records = len(df)
        date_range = f"{df['Date'].min()} to {df['Date'].max()}"
        
        # Calculate totals
        total_work_hours = df['Total_Work_Hours'].sum()
        total_overtime = df['Overtime_Hours'].sum()
        
        # Shift distribution
        day_shifts = len(df[df['Shift_Type'] == 'Day Shift'])
        night_shifts = len(df[df['Shift_Type'] == 'Night Shift'])
        
        return {
            'Metric': [
                'Total Records Processed',
                'Unique Employees',
                'Date Range',
                'Total Work Hours',
                'Total Overtime Hours',
                'Day Shift Entries',
                'Night Shift Entries',
                'Records with Overtime',
                'Average Hours per Entry',
                'Processing Date'
            ],
            'Value': [
                total_records,
                unique_employees,
                date_range,
                f"{total_work_hours:.2f}",
                f"{total_overtime:.2f}",
                day_shifts,
                night_shifts,
                len(df[df['Overtime_Hours'] > 0]),
                f"{total_work_hours/total_records:.2f}" if total_records > 0 else "0",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }

def create_premium_dashboard():
    """Create the premium dashboard interface"""
    
    # Premium Header
    st.markdown("""
    <div class="premium-header">
        <div class="premium-title">🏆 ULTIMATE TIMESHEET PROCESSOR</div>
        <div class="premium-subtitle">Premium Excel Enhancement System • Preserves Original Structure • Professional Results</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize processor
    processor = PremiumTimesheetProcessor()
    
    # Premium Sidebar
    with st.sidebar:
        st.markdown("""
        <div class="feature-card">
        <h2>🎛️ Premium Controls</h2>
        <p>Upload your timesheet file and enhance it with professional calculations</p>
        </div>
        """, unsafe_allow_html=True)
        
        # File Upload Section
        st.markdown("### 📂 File Upload")
        uploaded_file = st.file_uploader(
            "Choose your timesheet file",
            type=['csv', 'xlsx', 'xls'],
            help="Upload Excel or CSV timesheet files"
        )
        
        # Premium Features Display
        st.markdown("""
        <div class="feature-card">
        <h3>✨ Premium Features</h3>
        <ul>
        <li><strong>📋 Original Structure Preserved</strong></li>
        <li><strong>➕ Enhanced Calculations Added</strong></li>
        <li><strong>🎨 Professional Excel Formatting</strong></li>
        <li><strong>📊 Advanced Analytics</strong></li>
        <li><strong>⚡ Real-time Processing</strong></li>
        <li><strong>💎 Premium Visual Design</strong></li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
        
        # Business Rules
        st.markdown("""
        <div class="info-premium">
        <h3>🎯 Business Rules</h3>
        <p><strong>Day Shift:</strong> 8AM-5PM<br>
        Overtime after 5PM (30min-1.5h)</p>
        <p><strong>Night Shift:</strong> 6PM-3AM<br>
        Overtime after 3AM (30min-3h)</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Main Content Area
    if uploaded_file is not None:
        # Load and display file info
        st.markdown("## 📊 File Analysis & Enhancement")
        
        with st.spinner("🔄 Loading your timesheet data..."):
            raw_data = processor.load_timesheet_file_premium(uploaded_file)
        
        if raw_data is not None:
            # Premium File Overview
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{len(raw_data):,}</h1>
                <p>Total Records</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{raw_data['Name'].nunique()}</h1>
                <p>Unique Employees</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{len(raw_data['Date'].unique())}</h1>
                <p>Date Range (Days)</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{len(raw_data.columns)}</h1>
                <p>Data Columns</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Premium Sample Data Display
            st.markdown("### 📋 Original Data Preview")
            st.markdown("""
            <div class="info-premium">
            <strong>🔒 Original Structure Preserved:</strong> Your data order and employee names remain exactly as uploaded. 
            We only ADD enhancement columns - never modify or sort your original data.
            </div>
            """, unsafe_allow_html=True)
            
            st.dataframe(raw_data[['Name', 'Date', 'Time', 'Status']].head(10), width=None)
            
            # Duplicate Analysis with Premium Styling
            st.markdown("### 🔍 Data Pattern Analysis")
            
            duplicate_analysis = raw_data.groupby(['Name', 'Date']).size().reset_index(name='Entry_Count')
            multiple_entries = duplicate_analysis[duplicate_analysis['Entry_Count'] > 1]
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{len(duplicate_analysis):,}</h1>
                <p>Employee-Date Combinations</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{len(multiple_entries):,}</h1>
                <p>With Multiple Entries</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                duplicate_pct = len(multiple_entries)/len(duplicate_analysis)*100
                st.markdown(f"""
                <div class="metric-premium">
                <h1>{duplicate_pct:.1f}%</h1>
                <p>Requiring Consolidation</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Premium Entry Distribution Chart
            entry_distribution = duplicate_analysis['Entry_Count'].value_counts().sort_index()
            
            fig_dist = px.bar(
                x=entry_distribution.index,
                y=entry_distribution.values,
                title="📈 Entry Distribution Pattern",
                labels={'x': 'Entries per Employee per Day', 'y': 'Frequency'},
                color=entry_distribution.values,
                color_continuous_scale='Viridis'
            )
            fig_dist.update_layout(
                title_font_size=20,
                title_x=0.5,
                showlegend=False,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)'
            )
            st.plotly_chart(fig_dist, use_container_width=True)
            
            # Premium Processing Section
            st.markdown("## 🚀 Premium Data Enhancement")
            
            if st.button("🏆 ENHANCE WITH PREMIUM CALCULATIONS", type="primary", key="enhance_btn"):
                st.markdown("""
                <div class="success-premium">
                <h3>🚀 Starting Premium Enhancement Process</h3>
                <p>Processing your data with advanced algorithms while preserving original structure...</p>
                </div>
                """, unsafe_allow_html=True)
                
                with st.spinner("🔄 Applying premium calculations..."):
                    enhanced_data = processor.calculate_enhanced_data(raw_data)
                
                if enhanced_data is not None:
                    # Store in session state
                    st.session_state['enhanced_data'] = enhanced_data
                    st.session_state['raw_data'] = raw_data
                    
                    # Success Message
                    st.markdown("""
                    <div class="success-premium">
                    <h3>✅ PREMIUM ENHANCEMENT COMPLETED!</h3>
                    <p>Your data has been enhanced with professional calculations while maintaining original structure</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Enhancement Summary
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.markdown(f"""
                        <div class="metric-premium">
                        <h1>{len(enhanced_data.columns)}</h1>
                        <p>Total Columns</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col2:
                        new_cols = len(enhanced_data.columns) - len(raw_data.columns)
                        st.markdown(f"""
                        <div class="metric-premium">
                        <h1>+{new_cols}</h1>
                        <p>Enhanced Columns Added</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col3:
                        total_hours = enhanced_data['Total_Work_Hours'].sum()
                        st.markdown(f"""
                        <div class="metric-premium">
                        <h1>{total_hours:.0f}</h1>
                        <p>Total Work Hours</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with col4:
                        total_overtime = enhanced_data['Overtime_Hours'].sum()
                        st.markdown(f"""
                        <div class="metric-premium">
                        <h1>{total_overtime:.0f}</h1>
                        <p>Total Overtime Hours</p>
                        </div>
                        """, unsafe_allow_html=True)
    
    # Enhanced Results Section
    if 'enhanced_data' in st.session_state:
        enhanced_data = st.session_state['enhanced_data']
        
        st.markdown("## 🏆 Premium Enhanced Results")
        
        # Display enhanced data with highlighting
        st.markdown("### 📊 Enhanced Timesheet Data")
        st.markdown("""
        <div class="info-premium">
        <strong>🎯 Enhanced Columns Added:</strong> Consolidated_Start_Time, Consolidated_End_Time, Shift_Type, 
        Total_Work_Hours, Regular_Hours, Overtime_Hours, Entry_Count_For_Day
        </div>
        """, unsafe_allow_html=True)
        
        # Show sample of enhanced data
        display_columns = ['Name', 'Date', 'Time', 'Status', 'Consolidated_Start_Time', 
                          'Consolidated_End_Time', 'Shift_Type', 'Total_Work_Hours', 'Overtime_Hours']
        
        available_columns = [col for col in display_columns if col in enhanced_data.columns]
        st.dataframe(enhanced_data[available_columns].head(15), width=None)
        
        # Premium Analytics Section
        st.markdown("### 📈 Premium Analytics Dashboard")
        
        # Create premium analytics tabs
        tab1, tab2, tab3, tab4 = st.tabs(["🎯 Overview", "👥 Employee Analysis", "📅 Time Patterns", "💼 Overtime Details"])
        
        with tab1:
            col1, col2 = st.columns(2)
            
            with col1:
                # Shift Distribution
                shift_counts = enhanced_data['Shift_Type'].value_counts()
                fig_pie = px.pie(
                    values=shift_counts.values,
                    names=shift_counts.index,
                    title="🎯 Shift Type Distribution",
                    color_discrete_sequence=['#667eea', '#764ba2', '#f093fb']
                )
                fig_pie.update_layout(
                    title_font_size=18,
                    title_x=0.5
                )
                st.plotly_chart(fig_pie, use_container_width=True)
            
            with col2:
                # Work Hours Distribution
                fig_hist = px.histogram(
                    enhanced_data[enhanced_data['Total_Work_Hours'] > 0],
                    x='Total_Work_Hours',
                    title="⏰ Work Hours Distribution",
                    nbins=20,
                    color_discrete_sequence=['#4facfe']
                )
                fig_hist.update_layout(
                    title_font_size=18,
                    title_x=0.5,
                    xaxis_title="Hours Worked",
                    yaxis_title="Frequency"
                )
                st.plotly_chart(fig_hist, use_container_width=True)
        
        with tab2:
            # Employee Analysis
            employee_stats = enhanced_data.groupby('Name').agg({
                'Total_Work_Hours': 'sum',
                'Overtime_Hours': 'sum',
                'Date': 'nunique'
            }).reset_index()
            employee_stats.columns = ['Employee', 'Total Hours', 'Overtime Hours', 'Days Worked']
            employee_stats = employee_stats.sort_values('Total Hours', ascending=False)
            
            st.dataframe(employee_stats, width=None)
            
            # Top performers
            top_10 = employee_stats.head(10)
            fig_emp = px.bar(
                top_10,
                x='Employee',
                y='Total Hours',
                title="🏆 Top 10 Employees by Total Hours",
                color='Total Hours',
                color_continuous_scale='Viridis'
            )
            fig_emp.update_xaxes(tickangle=45)
            fig_emp.update_layout(
                title_font_size=18,
                title_x=0.5
            )
            st.plotly_chart(fig_emp, use_container_width=True)
        
        with tab3:
            # Time patterns
            enhanced_data['Date_parsed'] = pd.to_datetime(enhanced_data['Date'], format='%d/%m/%Y', errors='coerce')
            daily_stats = enhanced_data.groupby('Date_parsed').agg({
                'Total_Work_Hours': 'sum',
                'Overtime_Hours': 'sum',
                'Name': 'count'
            }).reset_index()
            daily_stats.columns = ['Date', 'Total Hours', 'Overtime Hours', 'Entries']
            
            fig_daily = px.line(
                daily_stats,
                x='Date',
                y='Total Hours',
                title="📅 Daily Total Hours Trend",
                markers=True,
                line_shape='spline'
            )
            fig_daily.update_layout(
                title_font_size=18,
                title_x=0.5
            )
            st.plotly_chart(fig_daily, use_container_width=True)
        
        with tab4:
            # Overtime analysis
            overtime_data = enhanced_data[enhanced_data['Overtime_Hours'] > 0]
            
            if not overtime_data.empty:
                col1, col2 = st.columns(2)
                
                with col1:
                    # Overtime distribution
                    fig_ot = px.histogram(
                        overtime_data,
                        x='Overtime_Hours',
                        title="💼 Overtime Hours Distribution",
                        nbins=15,
                        color_discrete_sequence=['#ff9500']
                    )
                    fig_ot.update_layout(
                        title_font_size=18,
                        title_x=0.5
                    )
                    st.plotly_chart(fig_ot, use_container_width=True)
                
                with col2:
                    # Overtime by shift
                    ot_by_shift = overtime_data.groupby('Shift_Type')['Overtime_Hours'].agg(['count', 'sum', 'mean']).reset_index()
                    ot_by_shift.columns = ['Shift Type', 'Count', 'Total OT', 'Average OT']
                    
                    fig_ot_shift = px.bar(
                        ot_by_shift,
                        x='Shift Type',
                        y='Total OT',
                        title="🌅 Overtime by Shift Type",
                        color='Shift Type',
                        color_discrete_sequence=['#667eea', '#764ba2']
                    )
                    fig_ot_shift.update_layout(
                        title_font_size=18,
                        title_x=0.5
                    )
                    st.plotly_chart(fig_ot_shift, use_container_width=True)
            else:
                st.info("📊 No overtime hours found in the data")
        
        # Premium Export Section
        st.markdown("## 💎 Premium Export")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # CSV Export
            csv_data = enhanced_data.to_csv(index=False)
            st.download_button(
                label="📄 Download Enhanced CSV",
                data=csv_data,
                file_name=f"enhanced_timesheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                type="primary"
            )
        
        with col2:
            # Premium Excel Export
            if st.button("📊 Generate Premium Excel", type="secondary"):
                with st.spinner("🎨 Creating premium Excel file..."):
                    excel_data, excel_filename = processor.export_premium_excel(enhanced_data)
                
                st.download_button(
                    label="💎 Download Premium Excel",
                    data=excel_data,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
                st.markdown("""
                <div class="success-premium">
                <h3>✨ Premium Excel Ready!</h3>
                <p>Professional formatting, highlighted enhanced columns, and summary sheets included</p>
                </div>
                """, unsafe_allow_html=True)
    
    # Premium Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                color: white; padding: 2rem; border-radius: 10px; margin-top: 2rem;">
    <h3>🏆 Ultimate Timesheet Processor - Premium Edition</h3>
    <p>Professional Excel Enhancement • Original Structure Preserved • Premium Visual Design</p>
    <p><strong>October 2025</strong> | Powered by Advanced Analytics</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    create_premium_dashboard()